#!/usr/bin/env python3
"""Smoke test for the Python `openpyxl` package.

Fully self-contained: no input files, no network, no packages beyond openpyxl
(and its implied deps). Exercises the core API surface and exits 0 only if every
check passes, so it can be used as a pass/fail library validator:

    python3 openpyxl.py

Install: pip install openpyxl   (import name: openpyxl)

This is the Python counterpart of data.table.R — same contract: a hard
not-installed guard (exit 1), a per-test harness that isolates failures, and a
PASS/exit-0 vs FAIL/exit-1 summary.
"""
import os
import sys

# This file is named after the package it tests, so it sits next to (and would
# shadow) the real top-level module. Drop this script's own directory from the
# import path before importing the package under test.
_here = os.path.dirname(os.path.abspath(__file__))
sys.path = [p for p in sys.path if p not in ("", ".") and os.path.abspath(p) != _here]

try:
    import openpyxl
except ImportError:
    print("FAIL: package 'openpyxl' is not installed")
    sys.exit(1)

import tempfile


def _version(mod, dist):
    """Best-effort version string: module.__version__, else installed metadata."""
    v = getattr(mod, "__version__", None)
    if v:
        return v
    try:
        import importlib.metadata as m

        return m.version(dist)
    except Exception:
        return "unknown"


print(f"openpyxl version: {_version(openpyxl, 'openpyxl')}")

failures = 0


def run_test(name, fn):
    """Run one check; a raised exception is a failure, not a crash."""
    global failures
    try:
        fn()
    except Exception as e:  # noqa: BLE001 - any failure is a test failure
        failures += 1
        print(f"  FAIL {name}: {e}")
    else:
        print(f"  ok   {name}")


def test_workbook_write_and_read_cells():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "x"
    ws["B2"] = 42
    assert ws["A1"].value == "x"
    assert ws["B2"].value == 42


def test_formula_cell_stored_verbatim():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 2
    ws["A2"] = 3
    ws["A3"] = "=SUM(A1:A2)"
    # openpyxl stores the formula string (it does not evaluate it).
    assert ws["A3"].value == "=SUM(A1:A2)"


def test_save_reopen_roundtrip_tempfile():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "x"
    ws["B2"] = 42
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        wb.save(path)
        assert os.path.getsize(path) > 0
        reopened = openpyxl.load_workbook(path)
        rws = reopened.active
        assert rws["A1"].value == "x"
        assert rws["B2"].value == 42
    finally:
        os.remove(path)


def test_append_rows_and_dimensions():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([1, 2, 3])
    ws.append([4, 5, 6])
    assert ws.max_row == 2
    assert ws.max_column == 3
    assert ws["C2"].value == 6


run_test("workbook write/read cells", test_workbook_write_and_read_cells)
run_test("formula cell stored verbatim", test_formula_cell_stored_verbatim)
run_test("save/reopen roundtrip tempfile", test_save_reopen_roundtrip_tempfile)
run_test("append rows + dimensions", test_append_rows_and_dimensions)

if failures > 0:
    print(f"FAIL: {failures} test(s) failed")
    sys.exit(1)
print("PASS: all openpyxl smoke tests passed")
